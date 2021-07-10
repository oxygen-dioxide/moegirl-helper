import yaml
from openpyxl import load_workbook
wb = load_workbook("moegirl-snippets.xlsx")
ws=wb[wb.sheetnames[0]]
snippet_dict={}
for i in range(2,ws.max_row):
    key=ws.cell(row=i,column=1).value
    if(key!=None and key[0]!="\\"):#以反斜杠开头的行视为注释
        prefixs=ws.cell(row=i,column=2).value.split(";")
        body="{{"+ws.cell(row=i,column=3).value
        if("\n" in body):
            body=body.split("\n")
        description=ws.cell(row=i,column=4).value
        #print(prefixs)
        prefixs=["{{"+i for i in prefixs]
        snippet_dict[key]={"prefix":prefixs,"body":body,"description":description}
outstr=""
line=""
with open("snippets.yaml",encoding="utf8") as infile:
    while(not(line.startswith("#="))):
        line=infile.readline()
        outstr+=line
outstr+=yaml.dump(snippet_dict,allow_unicode=True)
with open("snippets.yaml","w",encoding="utf8") as outfile:
    outfile.write(outstr)
    